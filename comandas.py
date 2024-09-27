
import openpyxl.workbook
import pandas as pd
from openpyxl.styles import Alignment
import os
from datetime import date

#Basic data
name='Kevin'
today_date=date.today()
delivery_to='CRYO INOX'
subgroup='INSTRUMENTACION'
offer='206021'
supplier='RTC (Siemens)'
contact_name='Òscar Sánchez'
contact_mail='oscar.sanchez@grupcarol.com/ndiaz@grupcarol.com'
contact_phone='(+34) 93 570 2644 / 682 623 236 '

# Load the excel table into a df. NEXT STEP, use data procesing to input the raw inst. list.
# include
pt_df=pd.read_excel('pt_list.xlsx')

wb=openpyxl.load_workbook('comandas_template.xlsx')
ws=wb.active

# Write constants
row_indx=5


for index,row in pt_df.iterrows():
    ws.cell(row=row_indx, column=1, value=name)
    ws.cell(row=row_indx, column=2, value=today_date)
    ws.cell(row=row_indx, column=3, value=delivery_to)
    ws.cell(row=row_indx, column=4, value=subgroup)
    ws.cell(row=row_indx, column=5, value=row['package_unit'])
    pressure_transmitter_info = f'''Transmisor de presion {row['model']} 
Segun oferta {offer} 
TAG: {row['tag']} 
Conexion: {row['process_connection']} 
Rango Calibracion: {row['cal_range']} 
Clasificacion ATEX: {row['atex_class']}'''
    
    # Write to column 7 and assign the cell to a variable
    cell = ws.cell(row=row_indx, column=7, value=pressure_transmitter_info)
    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    ws.cell(row=row_indx, column=9, value=1)
    ws.cell(row=row_indx, column=14, value=offer)
    ws.cell(row=row_indx, column=15, value=supplier)
    ws.cell(row=row_indx, column=18, value=row['price'])
    ws.cell(row=row_indx, column=19, value=row['lead_time'])
    ws.cell(row=row_indx, column=58, value=contact_name)
    ws.cell(row=row_indx, column=60, value=contact_mail)
    ws.cell(row=row_indx, column=59, value=contact_phone)
    

    row_indx+=1

wb.save('para_copiar_seguiment.xlsx')

print(f"Se ha generado el archivo {'para_copiar_seguiment.xlsx'} exitosamente, en el directorio {os.getcwd()}")






